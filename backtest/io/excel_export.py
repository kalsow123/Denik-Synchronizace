"""Excel export s více listy (grid výsledky + prop-firm)."""
from __future__ import annotations

from pathlib import Path
from typing import Any, Callable, Dict, FrozenSet, Optional

import pandas as pd

from backtest.grid.combo_columns import (
    BOT_NAME_COL,
    BOT_NAME_EXCEL_WIDTH_CM,
    COMBO_NO_COL,
)
from backtest.io.csv_export import _prepare_df_for_export
from backtest.prop_firm.compliance import PROP_FIRM_BOLD_COLUMNS

_CM_TO_EXCEL_COL_WIDTH = 4.724

GRID_REPORT_XLSX = "grid_report.xlsx"
GRID_SHEET_VYSLEDKY = "vysledky"
GRID_SHEET_SUMMARIES = "summaries"
GRID_SHEET_PROP_FIRM = "prop_firm"
GRID_SHEET_RANKING = "ranking"  # legacy název; nově Ranking_<PRESET>
GRID_SHEET_CHYBY = "chyby"
GRID_SHEET_DDI_EPIZODY = "ddi_epizody"
GRID_SHEET_E2E = "E2E"
RANKING_SHEET_PREFIX = "Ranking_"
_DDI_EPIZODE_VALUE_PREFIX = "DDi"
_LARGE_SHEET_FAST_THRESHOLD = 2000

# Tučné sloupce na listu Ranking_<PRESET> (backtest_risk_usd není tučně).
RANKING_BOLD_COLUMNS = frozenset({
    "headroom_scale",
    "max_risk_per_trade_usd",
    "projected_net_pnl_at_max_risk_usd",
    "original_net_pnl_usd",
    "max_dd_%_vs_initial",
    "profit_factor",
    "wave_min_pct",
})

COL_MAX_RISK = "max_risk_per_trade_usd"
COL_HEADROOM = "headroom_scale"
COL_PROJECTED_PNL = "projected_net_pnl_at_max_risk_usd"
COL_PROFIT_FACTOR = "profit_factor"
COL_WAVE_MIN_PCT = "wave_min_pct"
COL_MAX_DD_VS_INITIAL = "max_dd_%_vs_initial"
COL_RRR_TP = "RRR_TP"
# RRR_TP: wave_target_n_g končí „ G“ — tučně combo_no + RRR_TP ve všech listech
SUMMARY_RRR_TP_G_SUFFIX = " G"
_WAVE_TARGET_N_G_TP_MODES = frozenset({"wave_target_n_g", "wave_target_n_new"})

# List summaries — barvy textu (tučné)
SUMMARY_FONT_WAVE = "FF7B3F00"  # hnědá WAVE
SUMMARY_FONT_WAVE_COUNTER = "FF8E24AA"  # fialová WAVE_COUNTER
SUMMARY_FONT_WAVE_TWO_SIDED = "FF00332A"  # tmavě zelená WAVE_TWO_SIDED
SUMMARY_FONT_PP = "FF1B5E20"  # zelená PP
SUMMARY_FONT_EXT = "FF4682B4"  # modrá EXT
SUMMARY_FONT_BOS = "FFFF0000"  # červená BOS
SUMMARY_FONT_EXT_BOS = "FF000000"  # černá EXT_BOS

SUMMARY_COLS_WAVE = frozenset({
    "trades_wave",
    "net_pnl_wave_usd",
    "max_dd_%_vs_initial_wave",
})
SUMMARY_COLS_WAVE_COUNTER = frozenset({
    "trades_wave_counter",
    "net_pnl_wave_counter_usd",
    "max_dd_%_vs_initial_wave_counter",
})
SUMMARY_COLS_WAVE_TWO_SIDED = frozenset({
    "trades_wave_two_sided",
    "net_pnl_wave_two_sided_usd",
    "max_dd_%_vs_initial_wave_two_sided",
})
SUMMARY_COLS_PP = frozenset({
    "trades_pp",
    "net_pnl_pp_usd",
    "max_dd_%_vs_initial_pp",
})
SUMMARY_COLS_EXT = frozenset({
    "trades_ext",
    "net_pnl_ext_usd",
    "max_dd_%_vs_initial_ext",
})
SUMMARY_COLS_BOS = frozenset({
    "trades_bos",
    "net_pnl_bos_usd",
    "max_dd_%_vs_initial_bos",
})
SUMMARY_COLS_EXT_BOS = frozenset({
    "trades_ext_bos",
    "net_pnl_ext_bos_usd",
    "max_dd_%_vs_initial_ext_bos",
})

_FILL_PROJECTED_PNL = "D9D9D9"
_FILL_RED = "F8696B"
_FILL_YELLOW = "FFEB84"
_FILL_GREEN = "63BE7B"

# Jen pro zápis do .xlsx — interně zůstává tp_mode=wave_target_n_g (BotConfig, CSV, bot_name cache).
TP_MODE_XLSX_DISPLAY: dict[str, str] = {
    "wave_target_n_g": "wave_target_n_new",
}


def tp_mode_for_xlsx_display(value: Any) -> Any:
    """Přeloží tp_mode na popisek v Excelu (jen export, ne runtime)."""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return value
    raw = getattr(value, "value", value)
    key = str(raw).strip()
    return TP_MODE_XLSX_DISPLAY.get(key, raw)


def apply_xlsx_tp_mode_display_labels(df: pd.DataFrame) -> pd.DataFrame:
    """Před zápisem do .xlsx: wave_target_n_g → wave_target_n_new (+ bot_name segment)."""
    if df is None or df.empty:
        return df
    out = df.copy()
    if "tp_mode" in out.columns:
        out["tp_mode"] = out["tp_mode"].map(tp_mode_for_xlsx_display)
    if BOT_NAME_COL in out.columns:
        out[BOT_NAME_COL] = (
            out[BOT_NAME_COL]
            .astype(str)
            .str.replace("wave_target_n_g", "wave_target_n_new", regex=False)
        )
    return out


def is_ranking_sheet(sheet_name: str) -> bool:
    return str(sheet_name).startswith(RANKING_SHEET_PREFIX)


def _apply_sheet_layout(
    ws,
    df: pd.DataFrame,
    *,
    bot_name_width_scale: float = 1.0,
    bold_columns: Optional[FrozenSet[str]] = None,
    skip_row_bold: bool = False,
) -> None:
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    headers = list(df.columns)
    bold_set = bold_columns or frozenset()
    bold_font = Font(bold=True)
    bot_w = BOT_NAME_EXCEL_WIDTH_CM * _CM_TO_EXCEL_COL_WIDTH * bot_name_width_scale

    for col_idx, col_name in enumerate(headers, start=1):
        letter = get_column_letter(col_idx)
        if col_name == COMBO_NO_COL:
            ws.column_dimensions[letter].width = 10
        elif col_name == BOT_NAME_COL:
            ws.column_dimensions[letter].width = bot_w

    for col_idx, col_name in enumerate(headers, start=1):
        if col_name in bold_set:
            ws.cell(row=1, column=col_idx).font = bold_font

    if not skip_row_bold:
        for row_idx in range(2, ws.max_row + 1):
            for col_idx, col_name in enumerate(headers, start=1):
                if col_name in bold_set:
                    ws.cell(row=row_idx, column=col_idx).font = bold_font

    ws.freeze_panes = "B2"


def _interpolate_rgb(t: float, low: tuple[int, int, int], high: tuple[int, int, int]) -> str:
    t = max(0.0, min(1.0, float(t)))
    r = int(low[0] + (high[0] - low[0]) * t)
    g = int(low[1] + (high[1] - low[1]) * t)
    b = int(low[2] + (high[2] - low[2]) * t)
    return f"{r:02X}{g:02X}{b:02X}"


def _headroom_fill_hex(headroom: float) -> str:
    """1.0 = žlutá; <1 červená; >1 zelená (lineární přechod)."""
    h = max(0.0, float(headroom))
    red = (248, 105, 107)
    yellow = (255, 235, 132)
    green = (99, 190, 123)
    if h <= 1.0:
        if h <= 0.5:
            t = h / 0.5 if h > 0 else 0.0
            return _interpolate_rgb(t, (192, 0, 0), red)
        t = (h - 0.5) / 0.5
        return _interpolate_rgb(t, red, yellow)
    if h <= 2.0:
        t = (h - 1.0)
        return _interpolate_rgb(t, yellow, green)
    return _interpolate_rgb(1.0, yellow, green)


def _col_index(headers: list, name: str) -> Optional[int]:
    try:
        return headers.index(name) + 1
    except ValueError:
        return None


def _is_ddi_episode_value_column(col_name: str) -> bool:
    """Sloupce DDi1, DDi2, … (hodnota min. drawdownu v epizodě)."""
    name = str(col_name)
    if not name.startswith(_DDI_EPIZODE_VALUE_PREFIX):
        return False
    suffix = name[len(_DDI_EPIZODE_VALUE_PREFIX) :]
    return suffix.isdigit() and int(suffix) >= 1


def _ddi_episode_value_columns(headers: list) -> frozenset[str]:
    return frozenset(c for c in headers if _is_ddi_episode_value_column(c))


def _apply_ddi_epizody_sheet_format(ws, df: pd.DataFrame) -> None:
    """List ddi_epizody: tučně hodnoty ve sloupcích DDi1, DDi2, …"""
    from openpyxl.styles import Font

    headers = list(df.columns)
    bold = Font(bold=True)
    ddi_cols = _ddi_episode_value_columns(headers)

    for col_idx, col_name in enumerate(headers, start=1):
        if col_name not in ddi_cols:
            continue
        ws.cell(row=1, column=col_idx).font = bold
        for row_idx in range(2, ws.max_row + 1):
            ws.cell(row=row_idx, column=col_idx).font = bold


def _apply_ranking_sheet_format(ws, df: pd.DataFrame) -> None:
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    headers = list(df.columns)
    bold_font = Font(bold=True)
    gray_fill = PatternFill(start_color=_FILL_PROJECTED_PNL, end_color=_FILL_PROJECTED_PNL, fill_type="solid")

    idx_headroom = _col_index(headers, COL_HEADROOM)
    idx_max_risk = _col_index(headers, COL_MAX_RISK)
    idx_proj = _col_index(headers, COL_PROJECTED_PNL)
    idx_pf = _col_index(headers, COL_PROFIT_FACTOR)

    for row_idx in range(2, ws.max_row + 1):
        for col_idx, col_name in enumerate(headers, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if col_name in RANKING_BOLD_COLUMNS:
                cell.font = bold_font
            if col_name == COL_PROJECTED_PNL:
                cell.fill = gray_fill

        if idx_max_risk and idx_headroom:
            h_cell = ws.cell(row=row_idx, column=idx_headroom)
            try:
                h_val = float(h_cell.value) if h_cell.value not in (None, "") else 1.0
            except (TypeError, ValueError):
                h_val = 1.0
            fill_hex = _headroom_fill_hex(h_val)
            risk_cell = ws.cell(row=row_idx, column=idx_max_risk)
            risk_cell.fill = PatternFill(
                start_color=fill_hex, end_color=fill_hex, fill_type="solid"
            )
            risk_cell.font = bold_font

    if idx_pf and ws.max_row >= 2:
        letter = get_column_letter(idx_pf)
        rng = f"{letter}2:{letter}{ws.max_row}"
        ws.conditional_formatting.add(
            rng,
            ColorScaleRule(
                start_type="min",
                start_color=_FILL_RED,
                mid_type="percentile",
                mid_value=50,
                mid_color=_FILL_YELLOW,
                end_type="max",
                end_color=_FILL_GREEN,
            ),
        )
        for row_idx in range(2, ws.max_row + 1):
            ws.cell(row=row_idx, column=idx_pf).font = bold_font
        ws.cell(row=1, column=idx_pf).font = bold_font

    for col_idx, col_name in enumerate(headers, start=1):
        if col_name in RANKING_BOLD_COLUMNS:
            ws.cell(row=1, column=col_idx).font = bold_font


def _row_is_wave_target_n_g(df: pd.DataFrame, row_i: int) -> bool:
    """Řádek s tp_mode wave_target_n_g (nebo RRR_TP končící „ G“)."""
    if df is None or df.empty or row_i < 0 or row_i >= len(df):
        return False
    row = df.iloc[row_i]
    if COL_RRR_TP in df.columns:
        val = row[COL_RRR_TP]
        if isinstance(val, str) and val.endswith(SUMMARY_RRR_TP_G_SUFFIX):
            return True
    if "tp_mode" in df.columns:
        tm = row["tp_mode"]
        if tm is not None and not (isinstance(tm, float) and pd.isna(tm)):
            key = str(getattr(tm, "value", tm)).strip()
            if key in _WAVE_TARGET_N_G_TP_MODES:
                return True
    if BOT_NAME_COL in df.columns:
        bn = str(row[BOT_NAME_COL] or "")
        if "wave_target_n_g" in bn or "wave_target_n_new" in bn:
            return True
    return False


def _apply_wave_target_n_g_combo_rrr_bold(ws, df: pd.DataFrame) -> None:
    """Tučně combo_no a RRR_TP u řádků wave_target_n_g — všechny listy."""
    from openpyxl.styles import Font

    headers = list(df.columns)
    idx_combo = _col_index(headers, COMBO_NO_COL)
    idx_rrr = _col_index(headers, COL_RRR_TP)
    if not idx_combo and not idx_rrr:
        return

    bold = Font(bold=True)
    for row_idx in range(2, ws.max_row + 1):
        if not _row_is_wave_target_n_g(df, row_idx - 2):
            continue
        if idx_combo:
            ws.cell(row=row_idx, column=idx_combo).font = bold
        if idx_rrr:
            ws.cell(row=row_idx, column=idx_rrr).font = bold


def _apply_summaries_sheet_format(ws, df: pd.DataFrame) -> None:
    """
    List summaries: WAVE / WAVE_COUNTER / WAVE_TWO_SIDED / PP / BOS barevné tučné písmo;
    trades, profit_factor, max_dd_%_vs_initial (celkový), headroom, projected tučně;
    max_risk výplň dle headroom (stejná škála jako Ranking); profit_factor podmíněná škála min–max.
    """
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    headers = list(df.columns)
    bold = Font(bold=True)
    font_wave = Font(bold=True, color=SUMMARY_FONT_WAVE)
    font_wave_counter = Font(bold=True, color=SUMMARY_FONT_WAVE_COUNTER)
    font_wave_two_sided = Font(bold=True, color=SUMMARY_FONT_WAVE_TWO_SIDED)
    font_pp = Font(bold=True, color=SUMMARY_FONT_PP)
    font_ext = Font(bold=True, color=SUMMARY_FONT_EXT)
    font_bos = Font(bold=True, color=SUMMARY_FONT_BOS)
    font_ext_bos = Font(bold=True, color=SUMMARY_FONT_EXT_BOS)

    idx_headroom = _col_index(headers, COL_HEADROOM)
    idx_max_risk = _col_index(headers, COL_MAX_RISK)
    idx_pf = _col_index(headers, COL_PROFIT_FACTOR)
    def _hdr_font(col_name: str) -> Font:
        if col_name in SUMMARY_COLS_WAVE:
            return font_wave
        if col_name in SUMMARY_COLS_WAVE_COUNTER:
            return font_wave_counter
        if col_name in SUMMARY_COLS_WAVE_TWO_SIDED:
            return font_wave_two_sided
        if col_name in SUMMARY_COLS_PP:
            return font_pp
        if col_name in SUMMARY_COLS_EXT:
            return font_ext
        if col_name in SUMMARY_COLS_BOS:
            return font_bos
        if col_name in SUMMARY_COLS_EXT_BOS:
            return font_ext_bos
        if col_name in (
            COL_PROJECTED_PNL,
            COL_HEADROOM,
            COL_MAX_RISK,
            "trades",
            COL_PROFIT_FACTOR,
            COL_WAVE_MIN_PCT,
            COL_MAX_DD_VS_INITIAL,
        ):
            return bold
        return Font()

    for col_idx, col_name in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx).font = _hdr_font(col_name)

    for row_idx in range(2, ws.max_row + 1):
        for col_idx, col_name in enumerate(headers, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if col_name in SUMMARY_COLS_WAVE:
                cell.font = font_wave
            elif col_name in SUMMARY_COLS_WAVE_COUNTER:
                cell.font = font_wave_counter
            elif col_name in SUMMARY_COLS_WAVE_TWO_SIDED:
                cell.font = font_wave_two_sided
            elif col_name in SUMMARY_COLS_PP:
                cell.font = font_pp
            elif col_name in SUMMARY_COLS_EXT:
                cell.font = font_ext
            elif col_name in SUMMARY_COLS_BOS:
                cell.font = font_bos
            elif col_name in SUMMARY_COLS_EXT_BOS:
                cell.font = font_ext_bos
            elif col_name == COL_PROJECTED_PNL:
                cell.font = bold
            elif col_name == COL_HEADROOM:
                cell.font = bold
            elif col_name == COL_MAX_RISK:
                cell.font = bold
            elif col_name == "trades":
                cell.font = bold
            elif col_name == COL_PROFIT_FACTOR:
                cell.font = bold
            elif col_name == COL_WAVE_MIN_PCT:
                cell.font = bold
            elif col_name == COL_MAX_DD_VS_INITIAL:
                cell.font = bold

        if idx_max_risk and idx_headroom:
            h_cell = ws.cell(row=row_idx, column=idx_headroom)
            try:
                h_val = float(h_cell.value) if h_cell.value not in (None, "") else 1.0
            except (TypeError, ValueError):
                h_val = 1.0
            fill_hex = _headroom_fill_hex(h_val)
            risk_cell = ws.cell(row=row_idx, column=idx_max_risk)
            risk_cell.fill = PatternFill(
                start_color=fill_hex, end_color=fill_hex, fill_type="solid"
            )
            risk_cell.font = bold

    if idx_pf and ws.max_row >= 2:
        letter = get_column_letter(idx_pf)
        rng = f"{letter}2:{letter}{ws.max_row}"
        ws.conditional_formatting.add(
            rng,
            ColorScaleRule(
                start_type="min",
                start_color=_FILL_RED,
                mid_type="percentile",
                mid_value=50,
                mid_color=_FILL_YELLOW,
                end_type="max",
                end_color=_FILL_GREEN,
            ),
        )
        for row_idx in range(2, ws.max_row + 1):
            ws.cell(row=row_idx, column=idx_pf).font = bold
        ws.cell(row=1, column=idx_pf).font = bold


def export_grid_workbook(
    path: Path | str,
    sheets: Dict[str, pd.DataFrame],
    *,
    bot_name_width_cm: float = BOT_NAME_EXCEL_WIDTH_CM,
    bold_columns: Optional[FrozenSet[str]] = None,
    on_sheet_progress: Callable[[int, int, str], None] | None = None,
) -> bool:
    """
    Uloží jeden .xlsx soubor s více listy.
    Listy Ranking_<PRESET>: tučné sloupce, barevné škály dle specifikace.
    List summaries: WAVE/PP/BOS barvy textu, prop metriky, škály jako Ranking.
    """
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        return False

    width_factor = bot_name_width_cm * _CM_TO_EXCEL_COL_WIDTH / (
        BOT_NAME_EXCEL_WIDTH_CM * _CM_TO_EXCEL_COL_WIDTH
    )

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    def _write_sheet(sheet_name: str, df: pd.DataFrame) -> None:
        safe_name = str(sheet_name)[:31]
        prepared = _prepare_df_for_export(apply_xlsx_tp_mode_display_labels(df))
        prepared.to_excel(writer, sheet_name=safe_name, index=False)
        ws = writer.sheets[safe_name]
        large_sheet = len(prepared) > _LARGE_SHEET_FAST_THRESHOLD
        full_format = (
            is_ranking_sheet(safe_name)
            or safe_name in (GRID_SHEET_SUMMARIES, GRID_SHEET_E2E)
        )
        skip_row_bold = large_sheet and not full_format

        if is_ranking_sheet(safe_name):
            _apply_sheet_layout(
                ws,
                prepared,
                bot_name_width_scale=width_factor,
                bold_columns=frozenset(),
            )
            _apply_ranking_sheet_format(ws, prepared)
        elif safe_name in (GRID_SHEET_SUMMARIES, GRID_SHEET_E2E):
            _apply_sheet_layout(
                ws,
                prepared,
                bot_name_width_scale=width_factor,
                bold_columns=frozenset(),
            )
            _apply_summaries_sheet_format(ws, prepared)
        elif safe_name == GRID_SHEET_DDI_EPIZODY:
            _apply_sheet_layout(
                ws,
                prepared,
                bot_name_width_scale=width_factor,
                bold_columns=frozenset(),
            )
            _apply_ddi_epizody_sheet_format(ws, prepared)
        else:
            sheet_bold = bold_columns
            if sheet_name == GRID_SHEET_PROP_FIRM and sheet_bold is None:
                sheet_bold = PROP_FIRM_BOLD_COLUMNS
            _apply_sheet_layout(
                ws,
                prepared,
                bot_name_width_scale=width_factor,
                bold_columns=sheet_bold,
                skip_row_bold=skip_row_bold,
            )

        if not large_sheet or full_format:
            _apply_wave_target_n_g_combo_rrr_bold(ws, prepared)

    sheets_to_write = [
        (sheet_name, df)
        for sheet_name, df in sheets.items()
        if df is not None and not df.empty
    ]
    sheet_total = len(sheets_to_write) or 1

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        if sheets_to_write:
            for sheet_idx, (sheet_name, df) in enumerate(sheets_to_write, start=1):
                if on_sheet_progress is not None:
                    on_sheet_progress(sheet_idx, sheet_total, sheet_name)
                _write_sheet(sheet_name, df)
        else:
            # Init / checkpoint s prázdným reportem — openpyxl vyžaduje ≥1 viditelný list.
            if on_sheet_progress is not None:
                on_sheet_progress(1, 1, GRID_SHEET_VYSLEDKY)
            placeholder = pd.DataFrame(
                [
                    {
                        COMBO_NO_COL: None,
                        BOT_NAME_COL: "—",
                        "status": "(čeká na výsledky gridu)",
                    }
                ]
            )
            _write_sheet(GRID_SHEET_VYSLEDKY, placeholder)
    return True


def load_grid_report_sheet(path, sheet: str = GRID_SHEET_VYSLEDKY) -> pd.DataFrame:
    """Načte list z grid_report.xlsx."""
    return pd.read_excel(Path(path), sheet_name=sheet, engine="openpyxl")
