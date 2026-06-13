"""XLSX-only popisky tp_mode (wave_target_n_g → wave_target_n_new)."""

from __future__ import annotations

import pandas as pd
import pytest

from backtest.grid.summary_sheet import build_grid_summaries_sheet
from backtest.io.excel_export import (
    GRID_SHEET_SUMMARIES,
    GRID_SHEET_VYSLEDKY,
    apply_xlsx_tp_mode_display_labels,
    export_grid_workbook,
    load_grid_report_sheet,
    tp_mode_for_xlsx_display,
)
from config.enums import TPMode


def test_tp_mode_for_xlsx_display():
    assert tp_mode_for_xlsx_display("wave_target_n_g") == "wave_target_n_new"
    assert tp_mode_for_xlsx_display(TPMode.WAVE_TARGET_N_G) == "wave_target_n_new"
    assert tp_mode_for_xlsx_display("wave_target_n") == "wave_target_n"
    assert tp_mode_for_xlsx_display("rrr_fixed") == "rrr_fixed"


def test_apply_xlsx_labels_bot_name_and_tp_mode():
    df = pd.DataFrame(
        [
            {
                "combo_no": 1,
                "bot_name": "15_w0.3_o2_r2_f0.55_mkt_tpmwave_target_n_g_wave_extension_pct0.1_ASIA",
                "tp_mode": "wave_target_n_g",
            }
        ]
    )
    out = apply_xlsx_tp_mode_display_labels(df)
    assert out.iloc[0]["tp_mode"] == "wave_target_n_new"
    assert "wave_target_n_new" in out.iloc[0]["bot_name"]
    assert "wave_target_n_g" not in out.iloc[0]["bot_name"]


@pytest.mark.skipif(
    not pytest.importorskip("openpyxl", reason="openpyxl required"),
    reason="openpyxl",
)
def test_grid_workbook_xlsx_has_new_label(tmp_path):
    df = pd.DataFrame(
        [
            {
                "combo_no": 1,
                "bot_name": "bot_tpmwave_target_n_g",
                "tp_mode": "wave_target_n_g",
                "net_pnl_usd": 1.0,
            }
        ]
    )
    path = tmp_path / "grid_report.xlsx"
    assert export_grid_workbook(path, {GRID_SHEET_VYSLEDKY: df})
    loaded = load_grid_report_sheet(path)
    assert loaded.iloc[0]["tp_mode"] == "wave_target_n_new"
    assert "wave_target_n_new" in str(loaded.iloc[0]["bot_name"])


@pytest.mark.skipif(
    not pytest.importorskip("openpyxl", reason="openpyxl required"),
    reason="openpyxl",
)
def test_summaries_bolds_combo_no_and_rrr_tp_for_wave_target_n_g(tmp_path):
    df_report = pd.DataFrame(
        [
            {
                "combo_no": 7,
                "bot_name": "legacy",
                "timeframe": 15,
                "min_opp_bars": 2,
                "rrr": 2.0,
                "tp_mode": "wave_target_n",
                "fib_level": 0.5,
                "entry_mode": "market_fallback",
                "trades": 1,
                "profit_factor": 1.0,
                "max_dd_%_vs_initial": -1.0,
            },
            {
                "combo_no": 8,
                "bot_name": "g",
                "timeframe": 15,
                "min_opp_bars": 2,
                "rrr": 2.0,
                "tp_mode": "wave_target_n_g",
                "tp_target_wave_index": 4,
                "fib_level": 0.5,
                "entry_mode": "market_fallback",
                "trades": 2,
                "profit_factor": 1.1,
                "max_dd_%_vs_initial": -2.0,
            },
        ]
    )
    df_summaries = build_grid_summaries_sheet(df_report)
    path = tmp_path / "grid_report.xlsx"
    assert export_grid_workbook(
        path,
        {
            GRID_SHEET_VYSLEDKY: df_report,
            GRID_SHEET_SUMMARIES: df_summaries,
        },
    )

    from openpyxl import load_workbook

    wb = load_workbook(path)

    for sheet_name in (GRID_SHEET_VYSLEDKY, GRID_SHEET_SUMMARIES):
        ws = wb[sheet_name]
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        ix_combo = headers.index("combo_no") + 1
        assert ws.cell(2, ix_combo).font.bold is not True
        assert ws.cell(3, ix_combo).font.bold is True

    ws_sum = wb[GRID_SHEET_SUMMARIES]
    headers_sum = [ws_sum.cell(1, c).value for c in range(1, ws_sum.max_column + 1)]
    ix_rrr = headers_sum.index("RRR_TP") + 1
    assert ws_sum.cell(2, ix_rrr).font.bold is not True
    assert ws_sum.cell(3, ix_rrr).font.bold is True
    assert str(df_summaries.iloc[1]["RRR_TP"]).endswith(" G")
