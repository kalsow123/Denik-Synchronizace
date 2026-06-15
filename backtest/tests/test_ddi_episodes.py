"""DDi metriky, epizody drawdownu a list ddi_epizody."""

from __future__ import annotations

import pandas as pd
import pytest

from backtest.grid.ddi_sheet import build_grid_ddi_sheet
from backtest.metrics.dd_episodes import (
    find_dd_pct_vs_initial_episodes,
    format_dd_episodes_for_report,
    parse_dd_episodes_from_report,
)
from backtest.metrics.ddi_profile import build_daily_ddi_series, compute_ddi_profile
from backtest.stats import compute_stats


def _trades(rows: list[dict]) -> pd.DataFrame:
    df = pd.DataFrame(rows)
    df["close_time"] = pd.to_datetime(df["close_time"])
    return df


def test_wide_columns_two_episodes():
    results = {
        "bot_a": {
            "config": {"_grid_test_pozice": 1, "rrr": 2.5, "tp_mode": "rrr_fixed"},
            "total_trades": 10,
            "profit_factor": 1.1,
            "max_drawdown_pct": -12.0,
            "dd_episodes_ge10pct": [
                {
                    "start_date": "2024-01-02",
                    "end_recovery_date": "2024-01-05",
                    "min_dd_pct": -11.0,
                },
                {
                    "start_date": "2024-02-01",
                    "end_recovery_date": "2024-02-10",
                    "min_dd_pct": -10.5,
                },
            ],
            "ddi_profile": {"dnu_testu_celkem": 30, "pocet_epizod_ge10pct": 2},
        }
    }
    df_report = pd.DataFrame(
        [{"combo_no": 1, "trades": 10, "profit_factor": 1.1, "max_dd_%_vs_initial": -12.0}]
    )
    out = build_grid_ddi_sheet(results, df_report)
    assert "DD1_zacatek" in out.columns
    assert "DD1_konec" in out.columns
    assert "DDi1" in out.columns
    assert "DD2_zacatek" in out.columns
    assert "DD2_konec" in out.columns
    assert "DDi2" in out.columns
    assert out.loc[0, "DD1_zacatek"] == "2024-01-02"
    assert out.loc[0, "DDi1"] == -11.0
    assert out.loc[0, "DDi2"] == -10.5


def test_open_episode_konec_otevreno():
    results = {
        "bot_b": {
            "config": {"_grid_test_pozice": 2, "rrr": 2.0, "tp_mode": "wave_target_n", "tp_target_wave_index": 4},
            "total_trades": 5,
            "profit_factor": 0.9,
            "max_drawdown_pct": -14.1,
            "dd_episodes_ge10pct": [
                {"start_date": "2024-01-02", "end_recovery_date": "2024-01-04", "min_dd_pct": -11.0},
                {"start_date": "2025-04-28", "end_recovery_date": "OTEVRENO", "min_dd_pct": -14.1},
            ],
            "ddi_profile": {"dnu_testu_celkem": 100, "pocet_epizod_ge10pct": 2},
        }
    }
    out = build_grid_ddi_sheet(results, pd.DataFrame([{"combo_no": 2}]))
    assert out.loc[0, "DD2_konec"] == "OTEVRENO"
    assert out.loc[0, "DDi2"] == -14.1


def test_legacy_text_parses_ddi_values():
    legacy = "2024-08-28→2025-03-05 (-11.0%) | 2025-04-28→OTEVRENO (-14.1%)"
    results = {
        "bot_c": {
            "config": {"_grid_test_pozice": 3, "rrr": 2.5, "tp_mode": "rrr_fixed"},
            "total_trades": 1,
            "profit_factor": 1.0,
            "max_drawdown_pct": -14.1,
            "dd_ge_10pct_obdobi": legacy,
            "ddi_profile": {},
        }
    }
    out = build_grid_ddi_sheet(results, pd.DataFrame([{"combo_no": 3}]))
    assert out.loc[0, "DDi1"] == -11.0
    assert out.loc[0, "DDi2"] == -14.1
    parsed = parse_dd_episodes_from_report(legacy)
    assert len(parsed) == 2
    assert parsed[1]["end_recovery_date"] == "OTEVRENO"


def test_dnu_testu_celkem_fallback_from_dates():
    results = {
        "bot_d": {
            "config": {
                "_grid_test_pozice": 4,
                "date_from": "2026-01-01",
                "date_to": "2026-05-10",
                "rrr": 2.0,
                "tp_mode": "rrr_fixed",
            },
            "total_trades": 0,
            "profit_factor": 0,
            "max_drawdown_pct": 0,
            "ddi_profile": {},
        }
    }
    df_report = pd.DataFrame(
        [{"combo_no": 4, "date_from": "2026-01-01", "date_to": "2026-05-10"}]
    )
    out = build_grid_ddi_sheet(results, df_report)
    expected = (pd.Timestamp("2026-05-10") - pd.Timestamp("2026-01-01")).days + 1
    assert out.loc[0, "dnu_testu_celkem"] == expected


def test_episode_minus_13_pct_dates():
    df = _trades(
        [
            {"close_time": "2024-01-02", "pnl_usd": -5000.0, "close_reason": "SL"},
            {"close_time": "2024-01-03", "pnl_usd": -8000.0, "close_reason": "SL"},
            {"close_time": "2024-01-04", "pnl_usd": 13000.0, "close_reason": "TP"},
        ]
    )
    eps = find_dd_pct_vs_initial_episodes(df, initial_balance=100_000.0)
    assert len(eps) == 1
    ep = eps[0]
    assert ep["start_date"] == "2024-01-02"
    assert ep["trough_date"] == "2024-01-03"
    assert ep["end_recovery_date"] == "2024-01-04"
    assert ep["min_dd_pct"] == -13.0
    assert ep["is_open"] is False


def test_no_episode_when_max_dd_above_minus_10():
    df = _trades(
        [
            {"close_time": "2024-01-02", "pnl_usd": -3000.0, "close_reason": "SL"},
            {"close_time": "2024-01-03", "pnl_usd": 3000.0, "close_reason": "TP"},
        ]
    )
    eps = find_dd_pct_vs_initial_episodes(df, initial_balance=100_000.0)
    assert eps == []
    stats = compute_stats(df)
    assert stats["dd_episodes_ge10pct"] == []
    assert stats["dd_ge_10pct_obdobi"] == ""


def test_end_of_data_filtered_when_other_trades_exist():
    df = _trades(
        [
            {"close_time": "2024-01-02", "pnl_usd": -12000.0, "close_reason": "SL"},
            {"close_time": "2024-01-03", "pnl_usd": 0.0, "close_reason": "END_OF_DATA"},
        ]
    )
    eps = find_dd_pct_vs_initial_episodes(df)
    assert len(eps) == 1
    assert eps[0]["min_dd_pct"] == -12.0


def test_format_dd_episodes_roundtrip():
    eps = [
        {"start_date": "2024-08-28", "end_recovery_date": "2025-03-05", "min_dd_pct": -11.0},
        {"start_date": "2025-04-28", "end_recovery_date": "OTEVRENO", "min_dd_pct": -14.1},
    ]
    text = format_dd_episodes_for_report(eps)
    assert "2024-08-28→2025-03-05 (-11.0%)" in text
    assert "OTEVRENO (-14.1%)" in text


def test_compute_stats_includes_ddi_profile():
    df = _trades(
        [
            {"close_time": "2024-01-02", "pnl_usd": -12000.0, "close_reason": "SL"},
            {"close_time": "2024-01-10", "pnl_usd": 5000.0, "close_reason": "TP"},
        ]
    )
    stats = compute_stats(df)
    assert "ddi_profile" in stats
    assert "episodes" not in stats["ddi_profile"]
    assert len(stats["dd_episodes_ge10pct"]) == 1
    assert stats["ddi_profile"]["max_ddi_pct"] == pytest.approx(-12.0)
    assert stats["max_drawdown_pct"] == pytest.approx(-12.0)


def test_workbook_includes_ddi_epizody_sheet():
    from backtest.grid.grid_report_io import build_grid_workbook_sheets
    from backtest.io.excel_export import GRID_SHEET_DDI_EPIZODY

    results = {
        "bot_x": {
            "config": {
                "_grid_test_pozice": 1,
                "rrr": 2.5,
                "tp_mode": "wave_target_n",
                "tp_target_wave_index": 4,
                "date_from": "2026-01-01",
                "date_to": "2026-05-10",
            },
            "total_trades": 10,
            "profit_factor": 0.95,
            "max_drawdown_pct": -13.4,
            "dd_episodes_ge10pct": [
                {
                    "start_date": "2025-06-20",
                    "end_recovery_date": "OTEVRENO",
                    "min_dd_pct": -13.4,
                }
            ],
            "ddi_profile": {
                "dnu_testu_celkem": 151,
                "pocet_epizod_ge10pct": 1,
                "pct_dnu_ge_10": 13.25,
                "pct_dnu_v_dd": 94.04,
                "dnu_poruseni_10": 20,
                "median_ddi_pct": -4.87,
                "p90_ddi_pct": -10.26,
            },
        }
    }
    sheets, df_report, _df_long, _preset = build_grid_workbook_sheets(
        results, profile={"grid": []}, args=None
    )
    assert GRID_SHEET_DDI_EPIZODY in sheets
    ddi = sheets[GRID_SHEET_DDI_EPIZODY]
    assert ddi.loc[0, "combo_no"] == 1
    assert ddi.loc[0, "RRR_TP"] == "WAVE N=4"
    assert ddi.loc[0, "DD1_konec"] == "OTEVRENO"
    assert ddi.loc[0, "DDi1"] == -13.4


@pytest.mark.skipif(
    not pytest.importorskip("openpyxl", reason="openpyxl required"),
    reason="openpyxl",
)
def test_ddi_epizody_sheet_bolds_ddi_columns(tmp_path):
    from backtest.io.excel_export import GRID_SHEET_DDI_EPIZODY, export_grid_workbook

    df = pd.DataFrame(
        [
            {
                "combo_no": 1,
                "RRR_TP": "WAVE N=4",
                "trades": 10,
                "profit_factor": 0.95,
                "max_dd_%_vs_initial": -13.4,
                "dnu_testu_celkem": 151,
                "DD1_zacatek": "2025-06-20",
                "DD1_konec": "OTEVRENO",
                "DDi1": -13.4,
                "DD2_zacatek": "2025-08-01",
                "DD2_konec": "2025-09-01",
                "DDi2": -11.0,
                "pocet_epizod_ge10pct": 2,
            }
        ]
    )
    path = tmp_path / "grid_report.xlsx"
    assert export_grid_workbook(path, {GRID_SHEET_DDI_EPIZODY: df})

    from openpyxl import load_workbook

    ws = load_workbook(path)[GRID_SHEET_DDI_EPIZODY]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    ix_ddi1 = headers.index("DDi1") + 1
    ix_ddi2 = headers.index("DDi2") + 1
    ix_dd1_zac = headers.index("DD1_zacatek") + 1

    assert ws.cell(1, ix_ddi1).font.bold is True
    assert ws.cell(1, ix_ddi2).font.bold is True
    assert ws.cell(2, ix_ddi1).font.bold is True
    assert ws.cell(2, ix_ddi2).font.bold is True
    assert ws.cell(2, ix_dd1_zac).font.bold is not True


def test_daily_ddi_series_forward_fill():
    df = _trades(
        [
            {"close_time": "2024-01-02 15:00", "pnl_usd": -5000.0, "close_reason": "SL"},
            {"close_time": "2024-01-05 10:00", "pnl_usd": 1000.0, "close_reason": "TP"},
        ]
    )
    ddi = build_daily_ddi_series(
        df,
        initial_balance=100_000.0,
        date_from="2024-01-01",
        date_to="2024-01-05",
    )
    assert len(ddi) == 5
    assert ddi.iloc[0] == 0.0
    assert ddi.iloc[1] == pytest.approx(-5.0)
    assert ddi.iloc[2] == pytest.approx(-5.0)
    assert ddi.iloc[3] == pytest.approx(-5.0)
